import requests
import openpyxl
import json
from openpyxl import load_workbook

def call_api(method, url, body):
	"""
	Call an API with the given method, URL, and body.
	Returns status code and response body.
	"""
	try:
		# Parse body if it's a string (JSON)
		data = None
		if body and isinstance(body, str):
			try:
				data = json.loads(body)
			except json.JSONDecodeError:
				data = body
		
		# Make the API request
		if method.upper() == "GET":
			response = requests.get(url, params=data)
		elif method.upper() == "POST":
			response = requests.post(url, json=data)
		elif method.upper() == "PUT":
			response = requests.put(url, json=data)
		elif method.upper() == "DELETE":
			response = requests.delete(url, json=data)
		elif method.upper() == "PATCH":
			response = requests.patch(url, json=data)
		else:
			return None, f"Unsupported method: {method}"
		
		# Get response body
		try:
			response_body = response.json()
		except:
			response_body = response.text
		
		return response.status_code, json.dumps(response_body)
	
	except Exception as e:
		return None, str(e)

def main():
	# Get Excel file path from user
	excel_file = "4_call_api.xlsx"
	
	try:
		# Load the workbook
		wb = load_workbook(excel_file)
		ws = wb.active
		
		# Process each row (skip header row 1)
		for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
			# Get method, URL, and body from columns A, B, C
			method_cell = ws[f'A{row_idx}']
			url_cell = ws[f'B{row_idx}']
			body_cell = ws[f'C{row_idx}']
			
			method = method_cell.value
			url = url_cell.value
			body = body_cell.value
			
			# Skip empty rows
			if not method or not url:
				continue
			
			print(f"Processing row {row_idx}: {method} {url}")
			
			# Call the API
			status_code, response_body = call_api(method, url, body)
			
			# Write results to columns D and E
			status_cell = ws[f'D{row_idx}']
			response_cell = ws[f'E{row_idx}']
			
			status_cell.value = status_code
			response_cell.value = response_body
			
			print(f"  Status: {status_code}")
		
		# Save the workbook
		wb.save(excel_file)
		print(f"\nExcel file updated and saved: {excel_file}")
	
	except FileNotFoundError:
		print(f"Error: File '{excel_file}' not found.")
	except Exception as e:
		print(f"Error: {e}")

if __name__ == "__main__":
	main()
