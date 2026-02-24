import requests
import json
import threading
import time
from statistics import mean, stdev
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from collections import defaultdict
from datetime import datetime
import numpy as np

class PerformanceTest:
    def __init__(self, num_users=10, iterations=5, excel_file="4_call_api.xlsx"):
        """
        Initialize performance test
        
        Args:
            num_users: Number of concurrent users/threads
            iterations: Number of times each user calls each API
            excel_file: Path to Excel file containing API endpoints
        """
        self.num_users = num_users
        self.iterations = iterations
        self.excel_file = excel_file
        self.lock = threading.Lock()
        self.results = defaultdict(list)  # Store response times per API endpoint
        self.api_calls = []  # Store API details (method, URL, body)
        
    def call_api(self, method, url, body):
        """
        Call an API with the given method, URL, and body.
        Returns response time in milliseconds and status code.
        """
        try:
            data = None
            if body and isinstance(body, str):
                try:
                    data = json.loads(body)
                except json.JSONDecodeError:
                    data = body
            
            # Measure response time
            start_time = time.time()
            
            if method.upper() == "GET":
                response = requests.get(url, params=data, timeout=30)
            elif method.upper() == "POST":
                response = requests.post(url, json=data, timeout=30)
            elif method.upper() == "PUT":
                response = requests.put(url, json=data, timeout=30)
            elif method.upper() == "DELETE":
                response = requests.delete(url, json=data, timeout=30)
            elif method.upper() == "PATCH":
                response = requests.patch(url, json=data, timeout=30)
            else:
                return None, None
            
            elapsed_time = (time.time() - start_time) * 1000  # Convert to milliseconds
            return elapsed_time, response.status_code
        
        except requests.Timeout:
            return None, "TIMEOUT"
        except Exception as e:
            return None, f"ERROR: {str(e)}"
    
    def load_api_endpoints(self):
        """Load API endpoints from Excel file"""
        try:
            wb = load_workbook(self.excel_file)
            ws = wb.active
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
                method_cell = ws[f'A{row_idx}']
                url_cell = ws[f'B{row_idx}']
                body_cell = ws[f'C{row_idx}']
                
                method = method_cell.value
                url = url_cell.value
                body = body_cell.value
                
                if not method or not url:
                    continue
                
                self.api_calls.append({
                    'method': method,
                    'url': url,
                    'body': body,
                    'endpoint': f"{method.upper()} {url}"
                })
            
            print(f"Loaded {len(self.api_calls)} API endpoints from {self.excel_file}")
            return True
        
        except FileNotFoundError:
            print(f"Error: File '{self.excel_file}' not found.")
            return False
        except Exception as e:
            print(f"Error loading API endpoints: {e}")
            return False
    
    def worker(self, user_id):
        """
        Worker thread function - calls APIs multiple times
        
        Args:
            user_id: Identifier for this thread/user
        """
        for iteration in range(self.iterations):
            for api_call in self.api_calls:
                response_time, status_code = self.call_api(
                    api_call['method'],
                    api_call['url'],
                    api_call['body']
                )
                
                endpoint = api_call['endpoint']
                
                # Store results thread-safely
                with self.lock:
                    if response_time is not None:
                        self.results[endpoint].append(response_time)
                        print(f"User {user_id} | Iteration {iteration+1}/{self.iterations} | "
                              f"{endpoint} | Time: {response_time:.2f}ms | Status: {status_code}")
                    else:
                        print(f"User {user_id} | Iteration {iteration+1}/{self.iterations} | "
                              f"{endpoint} | FAILED | Status: {status_code}")
    
    def run_performance_test(self):
        """Run performance test with multiple threads"""
        print(f"\n{'='*80}")
        print(f"Starting Performance Test")
        print(f"Users: {self.num_users} | Iterations per user: {self.iterations}")
        print(f"Total API calls: {self.num_users * self.iterations * len(self.api_calls)}")
        print(f"{'='*80}\n")
        
        if not self.load_api_endpoints():
            return False
        
        # Create and start threads
        threads = []
        start_time = time.time()
        
        for user_id in range(1, self.num_users + 1):
            thread = threading.Thread(target=self.worker, args=(user_id,))
            threads.append(thread)
            thread.start()
        
        # Wait for all threads to complete
        for thread in threads:
            thread.join()
        
        elapsed_time = time.time() - start_time
        print(f"\n{'='*80}")
        print(f"Performance test completed in {elapsed_time:.2f} seconds")
        print(f"{'='*80}\n")
        
        return True
    
    def calculate_percentile(self, data, percentile):
        """Calculate percentile value"""
        if not data:
            return None
        return np.percentile(data, percentile)
    
    def generate_report(self):
        """Generate performance report and save to Excel"""
        print("Generating performance report...")
        
        # Create new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Performance Report"
        
        # Set up header
        headers = ["API Endpoint", "Total Calls", "Average (ms)", "Minimum (ms)", 
                   "Maximum (ms)", "90th Percentile (ms)", "Std Dev (ms)"]
        
        ws.append(headers)
        
        # Style header
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add data rows
        for endpoint in sorted(self.results.keys()):
            response_times = self.results[endpoint]
            
            if not response_times:
                continue
            
            total_calls = len(response_times)
            avg_time = mean(response_times)
            min_time = min(response_times)
            max_time = max(response_times)
            percentile_90 = self.calculate_percentile(response_times, 90)
            std_dev = stdev(response_times) if len(response_times) > 1 else 0
            
            ws.append([
                endpoint,
                total_calls,
                round(avg_time, 2),
                round(min_time, 2),
                round(max_time, 2),
                round(percentile_90, 2),
                round(std_dev, 2)
            ])
        
        # Set column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 15
        
        # Center align data
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Save report
        output_file = f"4_performance_report.xlsx"
        wb.save(output_file)
        
        print(f"\nPerformance report saved to: {output_file}")
        
        # Print summary to console
        print(f"\n{'='*80}")
        print("PERFORMANCE TEST SUMMARY")
        print(f"{'='*80}")
        for endpoint in sorted(self.results.keys()):
            response_times = self.results[endpoint]
            if response_times:
                print(f"\n{endpoint}")
                print(f"  Total Calls: {len(response_times)}")
                print(f"  Average: {mean(response_times):.2f} ms")
                print(f"  Minimum: {min(response_times):.2f} ms")
                print(f"  Maximum: {max(response_times):.2f} ms")
                print(f"  90th Percentile: {self.calculate_percentile(response_times, 90):.2f} ms")
                if len(response_times) > 1:
                    print(f"  Std Dev: {stdev(response_times):.2f} ms")
        print(f"\n{'='*80}\n")

def main():
    # Configuration
    NUM_USERS = 10  # Number of concurrent users
    ITERATIONS = 5  # Number of iterations per user
    EXCEL_FILE = "4_call_api.xlsx"  # Source Excel file with API endpoints
    
    # Create and run performance test
    perf_test = PerformanceTest(
        num_users=NUM_USERS,
        iterations=ITERATIONS,
        excel_file=EXCEL_FILE
    )
    
    # Run the test
    if perf_test.run_performance_test():
        perf_test.generate_report()
    else:
        print("Performance test failed!")

if __name__ == "__main__":
    main()
